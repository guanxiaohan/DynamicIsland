from Extension import * 
from typing import Self, Optional, List, Tuple, Dict, Any
from PySide6.QtCore import QTimer, QObject, QSize, Qt
from PySide6.QtWidgets import QMessageBox, QFileDialog
import os
import datetime
import dataclasses
import json
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import re


class ClassSchedulePanel(BarPanel):

    @dataclasses.dataclass
    class SingleClassTime:
        """Represents a single period/slot inside a TimeTable.  
        NOTE: repeat logic is removed from the slot - a TimeTable decides on which
        dates it applies (weekly / multiweekly / daily). A slot only defines when
        the period happens and its class_id + rendering/notification parameters.
        """

        @dataclasses.dataclass
        class TimeRule:
            # Minimal rule: start/end times + optional period index (1..n)
            start_time: datetime.time
            end_time: datetime.time
            period_index: int = 0

            def to_dict(self) -> dict:
                return {
                    "start_time": self.start_time.isoformat(),
                    "end_time": self.end_time.isoformat(),
                    "period_index": self.period_index,
                }

            @classmethod
            def from_dict(cls, d: dict) -> "ClassSchedulePanel.SingleClassTime.TimeRule":
                return cls(
                    start_time=datetime.time.fromisoformat(d["start_time"]),
                    end_time=datetime.time.fromisoformat(d["end_time"]),
                    period_index=int(d.get("period_index", 0)),
                )

        time_rule: "ClassSchedulePanel.SingleClassTime.TimeRule"
        class_id: int
        merged: bool = False
        ignored: bool = False
        display_name: str = "Class"
        notify_begin: int = 0
        notify_end: int = 0

        def dumpToJsonStr(self) -> str:
            data = {
                "time_rule": self.time_rule.to_dict(),
                "class_id": self.class_id,
                "merged": self.merged,
                "ignored": self.ignored,
                "display_name": self.display_name,
                "notify_begin": self.notify_begin,
                "notify_end": self.notify_end,
            }
            return json.dumps(data, ensure_ascii=False)

        @classmethod
        def loadFromJsonStr(cls, s: str) -> "ClassSchedulePanel.SingleClassTime":
            d = json.loads(s)
            return cls(
                time_rule=cls.TimeRule.from_dict(d["time_rule"]),
                class_id=int(d["class_id"]),
                merged=bool(d.get("merged", False)),
                ignored=bool(d.get("ignored", False)),
                display_name=d.get("display_name", "Class"),
                notify_begin=int(d.get("notify_begin", 0)),
                notify_end=int(d.get("notify_end", 0)),
            )

    @dataclasses.dataclass
    class ClassInstance:
        """Represents what class (name) a ClassID corresponds to on a weekday.
        NOTE: This no longer depends on TimeTable. It says: for ClassID X, on weekday W,
        the displayed class name should be Y. Expiry is supported for temporary changes.
        """
        class_name: str
        class_id: int
        weekday: int  # 1..7
        expire_on: Optional[datetime.datetime] = None

        def dumpToJsonStr(self) -> str:
            d = dataclasses.asdict(self)
            d["expire_on"] = self.expire_on.isoformat() if self.expire_on is not None else None
            return json.dumps(d, ensure_ascii=False)

        @classmethod
        def loadFromJsonStr(cls, s: str) -> "ClassSchedulePanel.ClassInstance":
            d = json.loads(s)
            exp = d.get("expire_on", None)
            exp_dt = datetime.datetime.fromisoformat(exp) if exp is not None else None
            return cls(class_name=d["class_name"], class_id=int(d["class_id"]), weekday=int(d["weekday"]), expire_on=exp_dt)

    @dataclasses.dataclass
    class TimeTableBundle:
        """A named TimeTable. It contains a list of slots (SingleClassTime).
        The bundle also describes when this TimeTable applies via cycle_mode & cycle_args.

        cycle_mode: Daily / Weekly / MultiWeekly
        cycle_args: free-text stored as written by user; parsed by applies_on_date.

        Multiweekly parsing convention (CycleArgs):
            "N;week1;week2;...;weekN"
        where weeki is a string of digits 1..7 indicating weekdays applicable in that week.
        Example: "4;123;;;67" means cycle_len=4, week1 Mondays-Tues-Weds, week2 none, week3 none, week4 Sat-Sun.
        """

        name: str
        cycle_mode: str  # "Daily", "Weekly", "MultiWeekly"
        cycle_args: str
        timeTable: List['ClassSchedulePanel.SingleClassTime']

        def dumpToJson(self) -> dict:
            return {
                "name": self.name,
                "cycle_mode": self.cycle_mode,
                "cycle_args": self.cycle_args,
                "timeTable": [t.dumpToJsonStr() for t in self.timeTable],
            }

        @classmethod
        def loadFromJson(cls, d: dict) -> 'ClassSchedulePanel.TimeTableBundle':
            return cls(
                name=d.get("name", "Unnamed"),
                cycle_mode=d.get("cycle_mode", "Weekly"),
                cycle_args=d.get("cycle_args", ""),
                timeTable=[ClassSchedulePanel.SingleClassTime.loadFromJsonStr(s) for s in d.get("timeTable", [])],
            )

        def applies_on_date(self, date: datetime.date) -> bool:
            """Decide whether this TimeTable should be used on a particular date.
            Interpretation:
            - Daily: always True
            - Weekly: cycle_args is a string of digits 1..7 (e.g. "12345") meaning weekdays
            - MultiWeekly: cycle_args = "N;w1;w2;...;wN" where each wi is digits for that week
            """
            mode = self.cycle_mode.lower() if isinstance(self.cycle_mode, str) else str(self.cycle_mode)
            if mode.startswith("d"):
                return True
            wd = date.isoweekday()
            if mode.startswith("w") and self.cycle_args:
                # weekly: digits indicate applicable weekdays
                try:
                    digits = tuple(int(ch) for ch in str(self.cycle_args) if ch.isdigit())
                    return wd in digits
                except Exception:
                    return True
            if mode.startswith("m"):
                parts = str(self.cycle_args).split(";") if self.cycle_args else []
                try:
                    cycle_len = int(parts[0]) if parts and parts[0].isdigit() else 0
                except Exception:
                    cycle_len = 0
                if cycle_len <= 0:
                    # fallback: behave as weekly
                    try:
                        digits = tuple(int(ch) for ch in str(self.cycle_args) if ch.isdigit())
                        return wd in digits
                    except Exception:
                        return True
                # determine week-in-cycle relative to a fixed epoch (Monday 1970-01-05)
                base_monday = datetime.date(1970, 1, 5)
                target_monday = date - datetime.timedelta(days=(date.weekday()))
                delta_weeks = ((target_monday - base_monday).days) // 7
                week_index = (delta_weeks % cycle_len) + 1
                # pick corresponding token
                if len(parts) >= week_index + 1:
                    token = parts[week_index]
                else:
                    token = parts[1] if len(parts) >= 2 else ""
                try:
                    digits = tuple(int(ch) for ch in token if ch.isdigit())
                    return wd in digits
                except Exception:
                    return False
            # default: apply
            return True

    class ClassSchedule:
        def __init__(self, timeTables: List['ClassSchedulePanel.TimeTableBundle'], classFills: List['ClassSchedulePanel.ClassInstance'], exceptions: List['ClassSchedulePanel.ClassInstance']):
            # timeTables: ordered list. For a given date we will check timeTables in order and gather slots from the first matching table
            self.timeTables = timeTables
            self.classes = classFills
            self.exceptions = exceptions

        def _find_applicable_timetable_for_date(self, date: datetime.date) -> Optional['ClassSchedulePanel.TimeTableBundle']:
            for tb in self.timeTables:
                try:
                    if tb.applies_on_date(date):
                        return tb
                except Exception:
                    continue
            return None

        def getCurrentClassInfo(self, ignore_merged_class: bool = True) -> dict:
            now = datetime.datetime.now()
            today = now.date()

            # build class map: (class_id, weekday) -> class_name (taking latest non-expired entry)
            class_map: Dict[Tuple[int, int], str] = {}
            for inst in self.classes:
                if inst.expire_on is not None and inst.expire_on <= now:
                    continue
                class_map[(inst.class_id, inst.weekday)] = inst.class_name

            # exceptions map by (date, class_id) -> new_name
            exceptions_map: Dict[Tuple[datetime.date, int], str] = {}
            for ex in self.exceptions:
                if ex.expire_on is not None and ex.expire_on <= now:
                    continue
                # for exceptions we treat weekday field as the target date (if stored that way) or use expire_on check
                # backwards-compat: if ex.weekday is > 31 assume it is actually encoded as date in class_name; but we store exceptions as date rows below
                if isinstance(ex.weekday, int) and 1 <= ex.weekday <= 31:
                    # nothing special - skip; this code path is for legacy rows; newer exceptions will be loaded with weekday representing target_date converted to ordinal
                    pass
            # For simplicity, load exceptions as separate structure: we assume exceptions list contains instances where class_name is new name and weekday stores date.toordinal()
            for ex in self.exceptions:
                try:
                    if ex.expire_on is not None and ex.expire_on <= now:
                        continue
                    # if weekday is stored as ordinal (>= 365) treat as date
                    if ex.weekday and ex.weekday > 366:
                        target_date = datetime.date.fromordinal(ex.weekday)
                        exceptions_map[(target_date, ex.class_id)] = ex.class_name
                except Exception:
                    continue

            LOOKAHEAD_DAYS = 28
            ongoing_candidates = []
            upcoming_candidates = []

            def make_datetimes_for_date(start_t: datetime.time, end_t: datetime.time, date_val: datetime.date) -> tuple[datetime.datetime, datetime.datetime]:
                st = datetime.datetime.combine(date_val, start_t)
                et = datetime.datetime.combine(date_val, end_t)
                if et <= st:
                    et = et + datetime.timedelta(days=1)
                return st, et

            for day_offset in range(0, LOOKAHEAD_DAYS):
                check_date = today + datetime.timedelta(days=day_offset)
                tb = self._find_applicable_timetable_for_date(check_date)
                if tb is None:
                    continue
                for slot in tb.timeTable:
                    if ignore_merged_class and getattr(slot, "merged", False):
                        continue
                    if getattr(slot, "ignored", False):
                        continue
                    tr = slot.time_rule
                    st, et = make_datetimes_for_date(tr.start_time, tr.end_time, check_date)
                    if st <= now < et:
                        ongoing_candidates.append((slot, st, et, tb))
                    elif st > now:
                        upcoming_candidates.append((slot, st, et, tb))
                if day_offset == 0 and ongoing_candidates:
                    break

            result = {
                "ongoing": False,
                "class_name": "",
                "class_id": None,
                "class_begin_time": None,
                "class_end_time": None,
            }

            def resolve_name_for_slot(slot: 'ClassSchedulePanel.SingleClassTime', date_of_slot: datetime.date) -> str:
                cid = slot.class_id
                # exceptions for that exact date
                if (date_of_slot, cid) in exceptions_map:
                    return exceptions_map[(date_of_slot, cid)]
                # else fallback to classes mapping (class_id, weekday)
                name = class_map.get((cid, date_of_slot.isoweekday()), "")
                return name

            if ongoing_candidates:
                chosen = min(ongoing_candidates, key=lambda t: t[2])
                slot, st, et, tb = chosen
                cid = slot.class_id
                name = resolve_name_for_slot(slot, st.date())
                result.update({
                    "ongoing": True,
                    "class_name": name,
                    "class_id": cid,
                    "class_begin_time": st,
                    "class_end_time": et,
                })
                return result

            if upcoming_candidates:
                chosen = min(upcoming_candidates, key=lambda t: t[1])
                slot, st, et, tb = chosen
                cid = slot.class_id
                name = resolve_name_for_slot(slot, st.date())
                result.update({
                    "ongoing": False,
                    "class_name": name,
                    "class_id": cid,
                    "class_begin_time": st,
                    "class_end_time": et,
                })
                return result

            return result

    PanelSizeHint = QSize(400, 30)

    def __init__(self):
        super().__init__()

        self.leftLabel = BasicLabel()
        self.rightLabel = BasicLabel()

        self.leftLayout.addWidget(self.leftLabel)
        self.rightLayout.addWidget(self.rightLabel, alignment=Qt.AlignmentFlag.AlignRight)

        self.updateTimer = QTimer()
        self.updateTimer.timeout.connect(self.updateDisplay)
        self.updateTimer.setInterval(4000)
        self.updateTimer.start()

    def postInitialize(self):
        self.loadSchedule()
        self.generateXlsxTemplate()

    def updateDisplay(self):
        if not hasattr(self, "scheduler") or not self.scheduler.classes:
            return

        def formatTimeToMins(secs: int):
            if secs < 60:
                return f"{secs}secs"
            if secs >= 86400:
                return f"{secs//86400}d {(secs%86400)//3600}h {(secs%3600)//60}min"
            if secs >= 3600:
                return f"{secs//3600}h {(secs%3600)//60}min"
            return f"{secs//60}min"

        info = self.scheduler.getCurrentClassInfo()
        begin_time: datetime.datetime = info['class_begin_time']
        if begin_time is None:
            self.requestHide.emit()
            return
        deltaSecs = (begin_time - datetime.datetime.now()).seconds

        MAXDELTA = 600
        if (not info["class_id"]) or (info['ongoing']) or deltaSecs > MAXDELTA:
            self.requestHide.emit()
            return

        self.leftLabel.transitionToText(
            f"Upcoming: {info['class_name']}"
        )
        self.rightLabel.transitionToText(
            f"In {formatTimeToMins(deltaSecs)} | {getTimeString(None, False)}"
        )
        self.requestShow.emit()

    def loadSchedule(self):
        # Backwards-compatible loader: loads single-file JSON layout if present
        schedules_path = ExtensionRoot + "ClassScheduler.Schedules.json"
        if os.path.exists(schedules_path):
            try:
                with open(schedules_path, encoding="utf-8") as f:
                    raw = json.load(f)
                timeTables = [self.TimeTableBundle.loadFromJson(x) for x in raw.get("timeTables", [])]
                classes = [self.ClassInstance.loadFromJsonStr(x) for x in raw.get("classes", [])]
                exceptions = [self.ClassInstance.loadFromJsonStr(x) for x in raw.get("exceptions", [])]
                self.scheduler = self.ClassSchedule(timeTables=timeTables, classFills=classes, exceptions=exceptions)
                # check for param inconsistencies across timetables (ignored/merged for same class_id)
                self._warn_param_inconsistencies()
                return
            except Exception as e:
                logger.info("Failed to load schedules.json", e)
        # if not found, keep empty
        self.scheduler = self.ClassSchedule(timeTables=[], classFills=[], exceptions=[])

    def _warn_param_inconsistencies(self):
        # Build map class_id -> list of param tuples from different timeTables
        param_map: Dict[int, List[Dict[str, Any]]] = {}
        for tb in self.scheduler.timeTables:
            for slot in tb.timeTable:
                cid = slot.class_id
                param = {"ignored": slot.ignored, "merged": slot.merged, "notify_begin": slot.notify_begin, "notify_end": slot.notify_end}
                param_map.setdefault(cid, []).append(param)
        for cid, plist in param_map.items():
            if len(plist) <= 1:
                continue
            base = plist[0]
            for other in plist[1:]:
                if other != base:
                    logger.info(f"Warning: class_id {cid} has differing slot-parameters across TimeTables. (e.g. {base} vs {other})")
                    break

    def saveSchedules(self):
        schedules_path = ExtensionRoot + "ClassScheduler.Schedules.json"
        try:
            data = {
                "timeTables": [tb.dumpToJson() for tb in self.scheduler.timeTables],
                "classes": [c.dumpToJsonStr() for c in self.scheduler.classes],
                "exceptions": [e.dumpToJsonStr() for e in self.scheduler.exceptions],
            }
            with open(schedules_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.info("Failed to save schedules:", e)

    @staticmethod
    def validate_timeTable_no_overlaps(timeTable: List["ClassSchedulePanel.SingleClassTime"]) -> None:
        # For the refactored model, we ensure slots within a single TimeTable do not overlap by time
        intervals: List[Tuple[int, int, "ClassSchedulePanel.SingleClassTime"]] = []
        classid_occurrence_count: Dict[int, int] = {}
        for slot in timeTable:
            start = slot.time_rule.start_time
            end = slot.time_rule.end_time
            start_minutes = start.hour * 60 + start.minute
            end_minutes = end.hour * 60 + end.minute
            if end_minutes > start_minutes:
                intervals.append((start_minutes, end_minutes, slot))
            else:
                intervals.append((start_minutes, 24 * 60, slot))
            classid_occurrence_count[slot.class_id] = classid_occurrence_count.get(slot.class_id, 0) + 1

        for cid, count in classid_occurrence_count.items():
            if count > 1:
                raise ValueError(f"Invalid TimeTable: class_id {cid} appears {count} times inside the same TimeTable.")

        intervals.sort(key=lambda t: (t[0], t[1]))
        for i in range(len(intervals) - 1):
            a_start, a_end, a_slot = intervals[i]
            b_start, b_end, b_slot = intervals[i + 1]
            if b_start < a_end:
                raise ValueError(f"TimeTable overlap detected: {a_slot.time_rule.start_time.strftime('%H:%M')}-{a_slot.time_rule.end_time.strftime('%H:%M')} (class_id={a_slot.class_id}) overlaps with {b_slot.time_rule.start_time.strftime('%H:%M')}-{b_slot.time_rule.end_time.strftime('%H:%M')} (class_id={b_slot.class_id})")

    def generateXlsxTemplate(self, save_to: str = ExtensionRoot + "ClassScheduler.Template.xlsx"):
        """
        New layout (refactored):
          - 'TimeTables' sheet: TableIndex | TableName | CycleMode (Daily/Weekly/MultiWeekly) | CycleArgs
          - For each timetable a sheet named 'TimeTable_<index>_<name>' with entries:
                PeriodIndex | ClassID | Class Role Name | Merged (Yes/No) | Ignored (Yes/No) | NotifyBegin | NotifyEnd | StartTime | EndTime
          - 'Classes' sheet: header: ClassID | Mon | Tue | Wed | Thu | Fri | Sat | Sun
          - 'Exceptions' sheet: Date (YYYY-MM-DD) | ClassID | NewClassName | ExpireOn (datetime optional)
        """
        wb = Workbook()
        ws_tt_list = wb.active
        if not ws_tt_list:
            return
        
        ws_tt_list.title = "TimeTables"
        ws_tt_list.append(["TableIndex", "TableName", "CycleMode (Daily/Weekly/MultiWeekly)", "CycleArgs (see docs)"])
        ws_tt_list.append([1, "Default", "Weekly", "12345"])
        ws_tt_list.append([2, "Evening", "Weekly", "67"])
        ws_tt_list.freeze_panes = "A2"
        for col_idx, width_val in {1:10,2:24,3:18,4:32}.items():
            ws_tt_list.column_dimensions[get_column_letter(col_idx)].width = width_val

        def create_time_table_sheet(title: str):
            ws = wb.create_sheet(title)
            headers = [
                "PeriodIndex",
                "ClassID",
                "Class Role Name",
                "Merged (Yes/No)",
                "Ignored (Yes/No)",
                "NotifyBegin (int)",
                "NotifyEnd (int)",
                "StartTime (time)",
                "EndTime (time)",
            ]
            ws.append(headers)
            ws.append([1, 11, "Physics", "No", "No", 0, 0, datetime.time(8,0), datetime.time(8,45)])
            ws.freeze_panes = "A2"
            dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            ws.add_data_validation(dv_yesno)
            dv_yesno.add("D2:D1000")
            dv_yesno.add("E2:E1000")
            dv_time = DataValidation(type="time", operator="between", formula1="TIME(0,0,0)", formula2="TIME(23,59,59)", allow_blank=False)
            ws.add_data_validation(dv_time)
            dv_time.add("H2:H1000")
            dv_time.add("I2:I1000")
            dv_int = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
            ws.add_data_validation(dv_int)
            dv_int.add("A2:A1000")
            dv_int.add("B2:B1000")
            for i,w in enumerate([10,10,28,10,10,12,12,12,12], start=1):
                ws.column_dimensions[get_column_letter(i)].width = w
            return ws

        create_time_table_sheet("TimeTable_1_Default")
        create_time_table_sheet("TimeTable_2_Evening")

        ws_cls = wb.create_sheet("Classes")
        header = ["ClassID", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        ws_cls.append(header)
        ws_cls.append([11, "Math", "Physics", "Chemistry", "Biology", "English", "", ""])
        ws_cls.freeze_panes = "A2"
        dv_int_cls = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=False)
        ws_cls.add_data_validation(dv_int_cls)
        dv_int_cls.add("A2:A1000")
        for i in range(1, len(header) + 1):
            ws_cls.column_dimensions[get_column_letter(i)].width = 16

        ws_ex = wb.create_sheet("Exceptions")
        ws_ex.append(["Date (YYYY-MM-DD)", "ClassID", "NewClassName", "ExpireOn (datetime optional)"])
        ws_ex.append([datetime.date(2025, 11, 20), 11, "GuestLecture", datetime.datetime(2025,11,20,16,0)])
        ws_ex.freeze_panes = "A2"
        dv_date_ex = DataValidation(type="date", operator="greaterThanOrEqual", formula1="DATE(1900,1,1)", allow_blank=False)
        ws_ex.add_data_validation(dv_date_ex)
        dv_date_ex.add("A2:A1000")
        dv_classid_ex = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=False)
        ws_ex.add_data_validation(dv_classid_ex)
        dv_classid_ex.add("B2:B1000")
        for i in range(1, 5):
            ws_ex.column_dimensions[get_column_letter(i)].width = 18

        wb.save(save_to)

    def dumpToXlsxFile(self, save_to: str = ExtensionRoot + "ClassScheduler.Dump.xlsx"):
        wb = Workbook()
        ws_list = wb.active
        if not ws_list:
            return
        
        ws_list.title = "TimeTables"
        ws_list.append(["TableIndex", "TableName", "CycleMode", "CycleArgs"])
        for idx, tb in enumerate(self.scheduler.timeTables, start=1):
            ws_list.append([idx, tb.name, tb.cycle_mode, tb.cycle_args])
        ws_list.freeze_panes = "A2"

        def add_time_table_sheet(idx: int, tb: ClassSchedulePanel.TimeTableBundle):
            title = f"TimeTable_{idx}_{tb.name}"
            ws = wb.create_sheet(title)
            headers = ["PeriodIndex", "ClassID", "Class Role Name", "Merged (Yes/No)", "Ignored (Yes/No)", "NotifyBegin", "NotifyEnd", "StartTime", "EndTime"]
            ws.append(headers)
            ws.freeze_panes = "A2"
            for slot in tb.timeTable:
                tr = slot.time_rule
                ws.append([
                    tr.period_index,
                    slot.class_id,
                    slot.display_name,
                    "Yes" if slot.merged else "No",
                    "Yes" if slot.ignored else "No",
                    slot.notify_begin,
                    slot.notify_end,
                    tr.start_time,
                    tr.end_time,
                ])
            if ws.max_row >= 2:
                ws["H2"].number_format = "HH:MM"
                ws["I2"].number_format = "HH:MM"
            return ws

        for idx, tb in enumerate(self.scheduler.timeTables, start=1):
            add_time_table_sheet(idx, tb)

        ws_cls = wb.create_sheet("Classes")
        ws_cls.append(["ClassID", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"])
        # gather class_ids
        class_rows: Dict[int, List[str]] = {}
        for ci in self.scheduler.classes:
            cid = ci.class_id
            if cid not in class_rows:
                class_rows[cid] = [""] * 7
            if 1 <= ci.weekday <= 7:
                class_rows[cid][ci.weekday - 1] = ci.class_name
        for cid, cells in class_rows.items():
            ws_cls.append([cid] + cells)
        ws_cls.freeze_panes = "A2"

        ws_ex = wb.create_sheet("Exceptions")
        ws_ex.append(["Date (YYYY-MM-DD)", "ClassID", "NewClassName", "ExpireOn (datetime optional)"])
        for ex in self.scheduler.exceptions:
            # if ex.weekday stored as ordinal, convert
            try:
                if ex.weekday and ex.weekday > 366:
                    date_val = datetime.date.fromordinal(ex.weekday)
                else:
                    date_val = None
            except Exception:
                date_val = None
            ws_ex.append([date_val, ex.class_id, ex.class_name, ex.expire_on if ex.expire_on is not None else None])
        ws_ex.freeze_panes = "A2"

        wb.save(save_to)

    def loadXlsxTemplate(self, load_from: str = ExtensionRoot + "ClassScheduler.Template.xlsx"):
        if not os.path.exists(load_from):
            raise FileNotFoundError(load_from)
        workbook = load_workbook(load_from, data_only=True)

        def parse_time_cell(value) -> Optional[datetime.time]:
            if value is None:
                return None
            if isinstance(value, datetime.time):
                return value
            if isinstance(value, datetime.datetime):
                return value.time()
            try:
                return datetime.time.fromisoformat(str(value))
            except Exception:
                try:
                    parts = str(value).strip().split(":")
                    hours = int(parts[0])
                    minutes = int(parts[1]) if len(parts) > 1 else 0
                    return datetime.time(hours, minutes)
                except Exception:
                    return None

        timeTables: List[ClassSchedulePanel.TimeTableBundle] = []
        class_fills_list: List[ClassSchedulePanel.ClassInstance] = []
        exceptions_list: List[ClassSchedulePanel.ClassInstance] = []

        if "TimeTables" in workbook.sheetnames:
            ws_list = workbook["TimeTables"]
            table_rows = list(ws_list.iter_rows(min_row=2, values_only=True))
            table_names = []
            for row in table_rows:
                if not row or row[1] is None:
                    continue
                tname = str(row[1])
                cycle_mode = str(row[2]) if row[2] is not None else "Weekly"
                cycle_args = str(row[3]) if len(row) > 3 and row[3] is not None else ""
                table_names.append((tname, cycle_mode, cycle_args))

            for idx, (tname, cycle_mode, cycle_args) in enumerate(table_names, start=1):
                sheet_name = f"TimeTable_{idx}_{tname}"
                tt_list: List[ClassSchedulePanel.SingleClassTime] = []
                if sheet_name in workbook.sheetnames:
                    ws_tt = workbook[sheet_name]
                    for row in ws_tt.iter_rows(min_row=2, values_only=True):
                        if not row or row[0] is None:
                            continue
                        try:
                            period_idx = int(row[0]) if row[0] is not None else 0 # type: ignore
                        except Exception:
                            period_idx = 0
                        try:
                            class_id_val = int(row[1]) if row[1] is not None else 0 # type: ignore
                        except Exception:
                            class_id_val = 0
                        display_name_val = str(row[2]) if row[2] is not None else "Class"
                        merged_val = str(row[3]).strip().lower() in ("yes", "true", "1", "y") if row[3] is not None else False
                        ignored_val = str(row[4]).strip().lower() in ("yes", "true", "1", "y") if row[4] is not None else False
                        notify_begin_val = int(row[5]) if row[5] is not None else 0 # type: ignore
                        notify_end_val = int(row[6]) if row[6] is not None else 0 # type: ignore
                        start_time_val = parse_time_cell(row[7])
                        end_time_val = parse_time_cell(row[8])
                        if start_time_val is None or end_time_val is None:
                            continue
                        time_rule_obj = ClassSchedulePanel.SingleClassTime.TimeRule(start_time=start_time_val, end_time=end_time_val, period_index=period_idx)
                        slot_obj = ClassSchedulePanel.SingleClassTime(time_rule=time_rule_obj, class_id=class_id_val, merged=merged_val, ignored=ignored_val, display_name=display_name_val, notify_begin=notify_begin_val, notify_end=notify_end_val)
                        tt_list.append(slot_obj)
                tb = ClassSchedulePanel.TimeTableBundle(name=tname, cycle_mode=cycle_mode, cycle_args=cycle_args, timeTable=tt_list)
                # validate single timetable
                self.validate_timeTable_no_overlaps(tb.timeTable)
                timeTables.append(tb)

            if "Classes" in workbook.sheetnames:
                ws_cls = workbook["Classes"]
                for row in ws_cls.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    try:
                        cid = int(row[0]) # type: ignore
                    except Exception:
                        continue
                    for offset in range(1,8):
                        val = row[offset] if offset < len(row) else None
                        if val is None or str(val).strip() == "":
                            continue
                        ci = ClassSchedulePanel.ClassInstance(class_name=str(val), class_id=cid, weekday=offset, expire_on=None)
                        class_fills_list.append(ci)

            if "Exceptions" in workbook.sheetnames:
                ws_ex = workbook["Exceptions"]
                now_dt = datetime.datetime.now()
                for row in ws_ex.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    try:
                        date_raw = row[0]
                        if isinstance(date_raw, datetime.datetime):
                            date_val = date_raw.date()
                        elif isinstance(date_raw, datetime.date):
                            date_val = date_raw
                        else:
                            date_val = datetime.date.fromisoformat(str(date_raw))
                        class_id_val = int(row[1]) if (len(row) > 1 and row[1] is not None) else None # type: ignore
                        new_name_val = str(row[2]) if (len(row) > 2 and row[2] is not None) else ""
                        expire_raw = row[3] if len(row) > 3 else None
                        expire_dt_val = None
                        if expire_raw:
                            if isinstance(expire_raw, datetime.datetime):
                                expire_dt_val = expire_raw
                            else:
                                expire_dt_val = datetime.datetime.fromisoformat(str(expire_raw))
                        if class_id_val is None:
                            continue
                        # we store the exception as ClassInstance where weekday stores date.toordinal() so that the schedule resolver can match exact dates
                        inst = ClassSchedulePanel.ClassInstance(class_name=new_name_val, class_id=class_id_val, weekday=date_val.toordinal(), expire_on=expire_dt_val)
                        if inst.expire_on is None or inst.expire_on > now_dt:
                            exceptions_list.append(inst)
                    except Exception:
                        continue

        else:
            # legacy handling (best-effort): try previous sheets and convert
            time_table_list = []
            class_fills_local = []
            exceptions_local = []
            if "TimeTable" in workbook.sheetnames:
                ws_tt = workbook["TimeTable"]
                for row in ws_tt.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    try:
                        class_id_val = int(row[0]) # type: ignore
                    except Exception:
                        continue
                    display_name_val = str(row[1]) if row[1] is not None else "Class"
                    merged_val = str(row[2]).strip().lower() in ("yes", "true", "1", "y") if row[2] is not None else False
                    ignored_val = str(row[3]).strip().lower() in ("yes", "true", "1", "y") if row[3] is not None else False
                    notify_begin_val = int(row[4]) if row[4] is not None else 0 # type: ignore
                    notify_end_val = int(row[5]) if row[5] is not None else 0 # type: ignore
                    start_time_val = parse_time_cell(row[6])
                    end_time_val = parse_time_cell(row[7])
                    if start_time_val is None or end_time_val is None:
                        continue
                    time_rule_obj = ClassSchedulePanel.SingleClassTime.TimeRule(start_time=start_time_val, end_time=end_time_val, period_index=0)
                    slot_obj = ClassSchedulePanel.SingleClassTime(time_rule=time_rule_obj, class_id=class_id_val, merged=merged_val, ignored=ignored_val, display_name=display_name_val, notify_begin=notify_begin_val, notify_end=notify_end_val)
                    time_table_list.append(slot_obj)

            if "Classes" in workbook.sheetnames:
                ws_cls = workbook["Classes"]
                for row in ws_cls.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    class_id_val = int(row[0]) # type: ignore
                    # try to map subsequent 7 cols to Mon..Sun
                    for idx_cell in range(1, min(8, len(row))):
                        cell_value = row[idx_cell]
                        if cell_value is None or str(cell_value).strip() == "":
                            continue
                        weekday_val = idx_cell
                        ci = ClassSchedulePanel.ClassInstance(class_name=str(cell_value), class_id=class_id_val, weekday=weekday_val, expire_on=None)
                        class_fills_local.append(ci)

            if "Exceptions" in workbook.sheetnames:
                ws_ex = workbook["Exceptions"]
                now_dt = datetime.datetime.now()
                for row in ws_ex.iter_rows(min_row=2, values_only=True):
                    if not row or row[0] is None:
                        continue
                    try:
                        date_raw = row[0]
                        if isinstance(date_raw, datetime.datetime):
                            date_val = date_raw.date()
                        elif isinstance(date_raw, datetime.date):
                            date_val = date_raw
                        else:
                            date_val = datetime.date.fromisoformat(str(date_raw))
                        class_id_val = int(row[1]) if (len(row) > 1 and row[1] is not None) else None # type: ignore
                        new_name_val = str(row[2]) if (len(row) > 2 and row[2] is not None) else ""
                        expire_raw = row[3] if len(row) > 3 else None
                        expire_dt_val = None
                        if expire_raw:
                            if isinstance(expire_raw, datetime.datetime):
                                expire_dt_val = expire_raw
                            else:
                                expire_dt_val = datetime.datetime.fromisoformat(str(expire_raw))
                        if class_id_val is None:
                            continue
                        inst = ClassSchedulePanel.ClassInstance(class_name=new_name_val, class_id=class_id_val, weekday=date_val.toordinal(), expire_on=expire_dt_val)
                        if inst.expire_on is None or inst.expire_on > now_dt:
                            exceptions_local.append(inst)
                    except Exception:
                        continue

            if time_table_list:
                self.validate_timeTable_no_overlaps(time_table_list)
                tb = ClassSchedulePanel.TimeTableBundle(name="Default", cycle_mode="Weekly", cycle_args="1234567", timeTable=time_table_list)
                timeTables.append(tb)
            class_fills_list.extend(class_fills_local)
            exceptions_list.extend(exceptions_local)

        self.scheduler = self.ClassSchedule(timeTables=timeTables, classFills=class_fills_list, exceptions=exceptions_list)
        # warn if slot params conflict between timetables
        self._warn_param_inconsistencies()
        self.saveSchedules()
        self.updateTimer.start()
        return

    def sysTrayItems(self):
        return {
            "Load from Template": self.loadDialog,
            "Save Schedule to Workbook": self.dumpDialog,
            "Generate a Empty Template": self.saveDialog,
        }

    def loadDialog(self):
        file, ok = QFileDialog.getOpenFileName(
            self, "Choose a template file", ".", "Excel 2010 Workbook (*.xlsx)"
        )
        if not file:
            return
        try:
            self.loadXlsxTemplate(file)
        except Exception as err:
            QMessageBox.critical(self, "Error", f"Failed to load the template.\n{err.__class__.__name__}: {err}")

    def saveDialog(self):
        file, ok = QFileDialog.getSaveFileName(
            self, "Generate a template", ExtensionRoot+"ClassScheduler.Template.xlsx", "Excel 2010 Workbook (*.xlsx)"
        )
        if not file:
            return
        try:
            self.generateXlsxTemplate(file)
        except Exception as err:
            QMessageBox.critical(self, "Error", f"Failed to generate the template.\n{err.__class__.__name__}: {err}")

    def dumpDialog(self):
        file, ok = QFileDialog.getSaveFileName(
            self, "Dump to a spreadsheet", ExtensionRoot+"ClassScheduler.Dump.xlsx", "Excel 2010 Workbook (*.xlsx)"
        )
        if not file:
            return
        try:
            self.dumpToXlsxFile(file)
        except Exception as err:
            QMessageBox.critical(self, "Error", f"Failed to load the template.\n{err.__class__.__name__}: {err}")


DI_setExtensionName("Class Scheduler")
DI_setExtensionNamespace("ClassScheduler")
DI_registerPanel("SchedulePanel", ClassSchedulePanel, 5)
